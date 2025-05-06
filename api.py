import helper
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import win32com.client
import time
variables_coord = None
results_coord = None

workbook_info = {
    "Sheet_Name": None,
    "Workbook": None,
    "Variables": [],
    "Results": [],
    "Variables_coord": None,
    "Results_coord": None,
    "Component_ID_coord": None,
    "Base_Unit_coord": None,
    "Base_Value_coord": None,
    "Values_coord": None,
    "Output_workbook_path": None,
    "Output_workbook": None,
}
def activate_workbook(path_to_workbook):
    """
    Activate the workbook at the specified path.
    
    :param path_to_workbook: Path to the workbook to activate.
    """

    try:
        # Attempt to load the workbook
        workbook = load_workbook(path_to_workbook, keep_vba=True)
    except PermissionError:
        print(f"Error: The file '{path_to_workbook}' is already open. Please close it and try again.")
        return None
    
    # Activate the workbook
    workbook.active = workbook.active
    
    workbook_info['Workbook'] = workbook

    base_name, ext = os.path.splitext(path_to_workbook)  # Split the file name and extension
    new_workbook_path = f"{base_name}-copy{ext}"         # Append '-copy' to the file name

    try:
        workbook.save(new_workbook_path)
    except PermissionError:
        print(f"Error: Unable to save the file '{new_workbook_path}'. It might be open in another program.")
        return None

    # Store the new workbook path in the global dictionary
    workbook_info['Output_workbook_path'] = new_workbook_path
    output_workbook = load_workbook(new_workbook_path, keep_vba=True)
    # Activate the output workbook
    output_workbook.active = output_workbook.active
    workbook_info['Output_workbook'] = output_workbook

    return workbook

def load_info(sheet_name): 
    if workbook_info['Workbook'] is None:
        raise ValueError("Workbook not set. Please load the workbook first.")
    workbook_info['Variables_coord'] = helper.find_cell_with_value(workbook_info['Workbook'], sheet_name, "Variables")
    workbook_info['Results_coord'] = helper.find_cell_with_value(workbook_info['Workbook'], sheet_name, "Results")
    workbook_info['Component_ID_coord'] = helper.find_cell_with_value(workbook_info['Workbook'], sheet_name, "Component ID")
    workbook_info['Base_Unit_coord'] = helper.find_cell_with_value(workbook_info['Workbook'], sheet_name, "Base Units")
    workbook_info['Base_Value_coord'] = helper.find_cell_with_value(workbook_info['Workbook'], sheet_name, "Base Value")
    workbook_info['Values_coord'] = helper.find_cell_with_value(workbook_info['Workbook'], sheet_name, "Value(s)...")
    workbook_info['Sheet_Name'] = sheet_name
    
def list_variables():
    """
    List all variables in the specified sheet.
    
    :param workbook: The workbook to search in.
    :param sheet_name: The name of the sheet to search in.
    :return: A list of variables found in the sheet.
    """
    
    if workbook_info['Workbook'] is None or workbook_info['Sheet_Name'] is None:
        raise ValueError("Workbook or Sheet Name not set. Please load the workbook and sheet name first.")
    
    # Get the row and column of the "Variables" cell
    start_row = workbook_info["Variables_coord"][0] + 2  # Start from the next row
    start_col = workbook_info["Variables_coord"][1]  # Column of the "Variables" cell
    
    
    # Iterate through rows until we reach an empty cell in the first column
    for row in range(start_row, workbook_info["Workbook"][workbook_info['Sheet_Name']].max_row + 1):
        cell_value = workbook_info["Workbook"][workbook_info['Sheet_Name']].cell(row=row, column=start_col).value
        if cell_value is None:
            break  # Stop if we reach an empty cell
        else:
            workbook_info['Variables'].append(cell_value)  # Add variable to the list
    
    return workbook_info['Variables']


def list_results():
    """
    List all Results in the specified sheet.
    
    :param workbook: The workbook to search in.
    :param sheet_name: The name of the sheet to search in.
    :return: A list of Results found in the sheet.
    """
    if workbook_info['Workbook'] is None or workbook_info['Sheet_Name'] is None:
        raise ValueError("Workbook or Sheet Name not set. Please load the workbook and sheet name first.")
    
    
    # Get the row and column of the "Results" cell
    start_row = workbook_info["Results_coord"][0] + 2  # Start from the next row
    start_col = workbook_info["Results_coord"][1]  # Column of the "Results" cell
    
    
    # Iterate through rows until we reach an empty cell in the first column
    for row in range(start_row, workbook_info["Workbook"][workbook_info['Sheet_Name']].max_row + 1):
        cell_value = workbook_info["Workbook"][workbook_info['Sheet_Name']].cell(row=row, column=start_col).value
        if cell_value is None:
            break  # Stop if we reach an empty cell
        else:
            workbook_info['Results'].append(cell_value)  # Add variable to the list
    
    return workbook_info['Results']


def set_value(var, value): 
    """
    Set the value of a variable in the specified sheet.
    
    :param workbook: The workbook to search in.
    :param sheet_name: The name of the sheet to search in.
    :param var: The variable to set the value for.
    :return: None
    """
    target_cell = helper.find_cell_with_value(workbook_info['Output_workbook'], workbook_info['Sheet_Name'], var)
    
    if target_cell: 
        row= target_cell[0]
        col= 8
        # Set the value of the cell to the new value
        sheet = workbook_info['Output_workbook'][workbook_info['Sheet_Name']]
        while sheet.cell(row=row, column=col).value is not None:
            col += 1
        sheet.cell(row=row, column=col).value = value
        workbook_info['Output_workbook'].save(workbook_info['Output_workbook_path'])

def run_study():
    time.sleep(10)
    """
    Add and run VBA code in the output workbook.
    """
    # Path to the output workbook
    output_workbook_path = workbook_info['Output_workbook_path']

    # Open Excel application
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # Optional: Set to True to make Excel visible

    # Open the output workbook
    workbook = excel.Workbooks.Open(output_workbook_path)

    # Add VBA code to the workbook
    vba_code = """
    Sub CallVSTOMethod()
        Dim addIn As COMAddIn
        Dim parametricStudy As Object
        Set addIn = Application.COMAddIns("VP.ExcelAddin")
        Set parametricStudy = addIn.Object
        parametricStudy.RunStudy
    End Sub
    """
    # Add the VBA code to a new module
    vb_module = workbook.VBProject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
    vb_module.CodeModule.AddFromString(vba_code)

    # Save the workbook with the VBA code
    workbook.Save()

    # Run the VBA macro
    excel.Application.Run("CallVSTOMethod")
    time.sleep(10)
    # Close the workbook and quit Excel
    workbook.Close(SaveChanges=True)
    excel.Quit()
    time.sleep(10)